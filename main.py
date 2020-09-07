# encoding='UTF-8'
import sys
import openpyxl  # excel操作模块
import configparser  # 导入配置.ini模块
from user_fun import style_excel


def main(argv=None):
    config = configparser.ConfigParser()  # 类实例化
    # 定义文件路径
    path = 'config.ini'
    config.read(path, encoding='UTF-8')
    inputFileName = config.get('select', 'inputFileName')
    outputFilename = config.get('select', 'outputFilename')
    学校名称 = config.get('select', 'schoolName')
    内地学生数 = config.getint('select', 'mainlandStu')
    港澳台学生数 = config.getint('select', 'hkMacaoTwStu')
    留学生 = config.getint('select', 'overseasStu')
    合计 = 内地学生数 + 港澳台学生数 + 留学生
    应到人数 = config.getint('select', 'arriveStu')
    # 读取表格
    wb = openpyxl.load_workbook(inputFileName)
    # 显示所有表单
    # 晨检异常
    晨检异常字典 = {"总数": 0, "体温异常": 0, "其他异常": 0, "详情": [["班级", "姓名", "是否体温异常", "晨检异常情况描述"]]}
    # 学生接触外省人员
    学生接触外省人员字典 = {"总数": 0, "详情": [["班级", "姓名", "文字描述"]]}
    # 家长接触重点疫区、境外
    家长接触重点疫区境外字典 = {"总数": 0, "详情": [["班级", "姓名", "文字描述"]]}
    # 请假学生
    请假学生字典 = {
        "总数": 0,
        "因发热未到或请假": 0,
        "因咳嗽腹泄乏力请假人数": 0,
        "因其他原因请假人数": 0,
        "详情": [["班级", "姓名", "是否发热", "是否咳嗽乏力腹泻呕吐等", "是否有重点疫区接触史", "请假原因"]]
    }
    班级 = ""
    晨检午检异常描述 = ""
    乘坐公交车学生人数 = 0
    for cell in wb["主表"]["J"]:
        if (cell.row > 2):
            乘坐公交车学生人数 += int(cell.value)

    for cell in wb["晨检异常登记"]["C"]:
        if (cell.value != "") and (cell.value != "姓名"):
            晨检异常字典["总数"] += 1
            if (wb["晨检异常登记"].cell(cell.row, 4).value != "否") and (wb["请假学生登记"].cell(
                    cell.row, 4).value != "无"):
                晨检异常字典["体温异常"] += 1
            for it in wb["主表"]["R"]:
                if (str(it.value).find(cell.value) >= 0):
                    班级 = str(wb["主表"].cell(it.row, 6).value) + \
                        str(wb["主表"].cell(it.row, 7).value)
            晨检午检异常描述 += str(班级 + cell.value + wb["晨检异常登记"].cell(cell.row, 5).value) + "\n"
            晨检异常字典["详情"].append([
                班级, cell.value, wb["晨检异常登记"].cell(cell.row, 4).value,
                wb["晨检异常登记"].cell(cell.row, 5).value
            ])
    for cell in wb["学生接触外省人员登记"]["C"]:
        if (cell.value != "") and (cell.value != "姓名"):
            学生接触外省人员字典["总数"] += 1
            for it in wb["主表"]["R"]:
                if (str(it.value).find(cell.value) >= 0):
                    班级 = str(wb["主表"].cell(it.row, 6).value) +\
                        str(wb["主表"].cell(it.row, 7).value)
            学生接触外省人员字典["详情"].append([班级, cell.value, wb["学生接触外省人员登记"].cell(cell.row, 4).value])
    for cell in wb["家长接触重点疫区、境外登记"]["C"]:
        if (cell.value != "") and (cell.value != "姓名"):
            家长接触重点疫区境外字典["总数"] += 1
            for it in wb["主表"]["R"]:
                if (str(it.value).find(cell.value) >= 0):
                    班级 = str(wb["主表"].cell(it.row, 6).value) +\
                        str(wb["主表"].cell(it.row, 7).value)
            家长接触重点疫区境外字典["详情"].append([班级, cell.value, wb["家长接触重点疫区、境外登记"].cell(cell.row, 4).value])
    for cell in wb["请假学生登记"]["C"]:
        if (cell.value != "") and (cell.value != "姓名"):
            请假学生字典["总数"] += 1
            if (wb["请假学生登记"].cell(cell.row, 4).value != "否") and (wb["请假学生登记"].cell(
                    cell.row, 4).value != "无"):
                请假学生字典["因发热未到或请假"] += 1
            if (wb["请假学生登记"].cell(cell.row, 5).value != "否") and (wb["请假学生登记"].cell(
                    cell.row, 5).value != "无"):
                请假学生字典["因咳嗽腹泄乏力请假人数"] += 1
            for it in wb["主表"]["R"]:
                if (str(it.value).find(cell.value) >= 0):
                    班级 = str(wb["主表"].cell(it.row, 6).value) +\
                        str(wb["主表"].cell(it.row, 7).value)
            请假学生字典["详情"].append([
                班级, cell.value, wb["请假学生登记"].cell(cell.row,
                                                  4).value, wb["请假学生登记"].cell(cell.row, 5).value,
                wb["请假学生登记"].cell(cell.row, 6).value, wb["请假学生登记"].cell(cell.row, 7).value
            ])
        else:
            wb["请假学生登记"].delete_rows(cell.row)
    晨检异常字典["其他异常"] = 晨检异常字典["总数"] - 晨检异常字典["体温异常"]
    请假学生字典["因其他原因请假人数"] = 请假学生字典["总数"] - 请假学生字典["因发热未到或请假"] - 请假学生字典["因咳嗽腹泄乏力请假人数"]
    # EXCEL 新表
    wbNEW = openpyxl.Workbook()
    wbNEW.active.title = "晨检异常"
    wbNEW.create_sheet("学生接触外省人员")
    wbNEW.create_sheet("家长接触重点疫区、境外")
    wbNEW.create_sheet("请假学生")
    wbNEW.create_sheet("综合统计")
    for cell in 晨检异常字典["详情"]:
        wbNEW["晨检异常"].append(cell)
    for cell in 学生接触外省人员字典["详情"]:
        wbNEW["学生接触外省人员"].append(cell)
    for cell in 家长接触重点疫区境外字典["详情"]:
        wbNEW["家长接触重点疫区、境外"].append(cell)
    for cell in 请假学生字典["详情"]:
        wbNEW["请假学生"].append(cell)
    wbNEW["综合统计"].append([
        "学校名称", "应到人数(小学)","实到人数(小学)", "未报到总数","因发热未报到人数","因其他原因未报到人数",\
        "请假总数(小学)", "因发热请假（体温37.3以上）", "因咳嗽、腹泄、乏力请假人数", "因其他原因请假人数", \
        "晨午检异常人数（体温37.3以上）","晨午检异常其他人数（如：咳嗽、腹泻等）：", "晨检午检异常描述",\
        "接触过疫情中高风险地区返衢对象的(学生数)","接触过境外返衢对象的(学生数)","乘坐公交学生数",\
        "今日到校，港澳台学生几个", "今日到校，外籍学生几个", "与疫情中高风险地区、境外回衢人员接触的家长人数统计",
    ])
    wbNEW["综合统计"].append([
        学校名称,应到人数,应到人数 - 请假学生字典["总数"],\
        0,0,0, 0,\
        请假学生字典["总数"],请假学生字典["因发热未到或请假"],请假学生字典["因咳嗽腹泄乏力请假人数"],请假学生字典["因其他原因请假人数"],\
        晨检异常字典["体温异常"],晨检异常字典["其他异常"],晨检异常字典["晨检午检异常描述"],\
        学生接触外省人员字典["总数"],0,乘坐公交车学生人数,0,1,家长接触重点疫区境外字典["总数"],
        ])
    # 设置自动列宽
    style_excel(self=None, target_wb=wbNEW, sheet_name="综合统计")
    wbNEW.save(outputFilename)


if __name__ == "__main__":
    sys.exit(main())
